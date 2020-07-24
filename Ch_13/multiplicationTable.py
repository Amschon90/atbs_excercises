# Create a program multiplicationTable.py that takes a number N from the command line
# and creates an N×N multiplication table in an Excel spreadsheet.

import os, openpyxl, sys
from openpyxl.styles import Font

# Set directory, bring in myNum..
os.chdir('C:\\Users\\Zig0n\\Documents\\VSCode Projects\\atbs_excercises\\Ch_13')
myNum = int(sys.argv[1])

# Open blank file and select sheet
wb = openpyxl.Workbook()
sheet = wb.active

# Create header and column row in bold
for num in range(1,myNum+1):
    fontObj = Font(bold=True)
    sheet.cell(row=1,column=num+1).font = fontObj
    sheet.cell(row=num+1,column=1).font = fontObj
    sheet.cell(row=1,column=num+1).value = num
    sheet.cell(row=num+1,column=1).value = num

# Create table values
for c in range(1,sheet.max_column):
    for r in range(1,sheet.max_row):
        sheet.cell(row=r+1,column=c+1).value = c*r

# Save the spreadsheet
wb.save(f'multiplicationTable_{myNum}.xlsx')
print(f'multiplicationTable_{myNum}.xlsx created!')